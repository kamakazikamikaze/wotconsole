#!/bin/env python
from __future__ import print_function
import os
import sys
import openpyxl
from ciscoconfparse import CiscoConfParse
import warnings
import re
from natsort import natsorted
from itertools import groupby
from operator import itemgetter
import tempfile


try:  # Portability for Python 3, though 2to3 would convert it
    input = raw_input
    xrange = range
except NameError:
    pass

# This is a simplified set of models and cannot be expected to cover every
# case/scenario. Feeds on a 3560X will not match those on a 3560(CX/G). For
# this reason, the option to override the "legal" feed ports is now available
feed_ports_regex = {
    '3560': r'Gi?0/11$|Gi?0/12$',
    '3750': r'Gi?[1-7]/0/[1-4]$',
    '3850': r'Te?[1-7]/1/[1-4]$|Gi?[1-7]/1/[1-4]$',  # Gig OR TenGig
    '4506': r'Te?[1-6]/[1-2]$|Gi?[1-6]/[3-6]$'
}

switch_models = ['3560', '3750', '3850', '4506', 'Other']

cutsheet_dir = './cutsheets/'

config_dir = './configs/'

output_dir = './output/'

template_dir = './templates/'


def condensify_ports(ports):
    """
    Turn a collection of ports into a Cisco-formatted range

    .. todo:: Altering format depending on "new" switch model

    :param ports: List of port names

    :returns: Ports in a Cisco-formatted range
    :rtype: String
    """
    keysplit = re.compile(r'[FGT](\d/)+')
    groups = {}
    for port in ports:
        try:
            key = [m.span() for m in keysplit.finditer(port)][0]
        except:
            continue
        try:
            groups[port[key[0]:key[1]]].add(int(port[key[1]:]))
        except KeyError:
            groups[port[key[0]:key[1]]] = set([int(port[key[1]:])])
        except:
            continue
    grouped = {}
    for key in groups.keys():
        for k, g in groupby(
                enumerate(sorted(groups[key])), lambda ix: ix[0] - ix[1]):
            try:
                grouped[key].append(map(itemgetter(1), g))
            except KeyError:
                grouped[key] = [map(itemgetter(1), g)]
    start = True
    condensed = ''
    for key, value in sorted(grouped.iteritems()):
        for v in value:
            if not start:
                condensed += ', '
            condensed += key + str(v[0])
            if len(v) > 1:
                condensed += '-' + str(v[-1])
            start = False
    return condensed


def get_vlan_list(oldconfig, regex):
    """
    Retrieve all VLANs from the old configuration file and return a list.

    (This is intended for future use in cross-platform conversions.)

    :param oldconfig: CiscoConfParse object of existing configuration file
    :param regex: Regex string used to determine if port is a feed

    :returns: All VLANs defined, sorted in ascending order
    :rtype: List
    """
    vlans = []
    allvlans = oldconfig.find_objects(r"^vlan \d+")
    allvlans.reverse()
    for obj in allvlans:
        # vlans.append(obj.text.split()[-1])
        vlan = obj.text.split()[-1]
        vlans.append(vlan)
    return natsorted(vlans, key=lambda x: x.lower())


def vlan_extract(oldconfig, newconfig, regex, genconfig=False):
    """
    Retrieve all VLANs from the old configuration file

    Automatically detects if certain VLANs will not be used and will offer to
    prune them.

    .. note:: For pruning to work, this _must_ be called before
              :py:func:`setup_feeds`

    :param oldconfig: CiscoConfParse object of existing configuration file
    :param newconfig: CiscoConfParse object, representing the "new" config file
                 defaults to None
    :param regex: Regex string used to determine if port is a feed
    :param genconfig: A boolean representing if a full config will be
                      generated:
                      If True, all VLANs will be added to the new config file.
                      Defaults to False

    :returns: All VLANs defined, sorted in ascending order
    :rtype: List
    """
    vlans = []
    allvlans = oldconfig.find_objects(r"^vlan \d+")
    allvlans.reverse()
    for obj in allvlans:
        # vlans.append(obj.text.split()[-1])
        vlan = obj.text.split()[-1]
        usedby = newconfig.find_objects_w_child(
            r"^interface [FfGgTt]\d", r"vlan.*" + vlan)
        usedby = [u for u in usedby if not u.re_match(regex, 0)]
        vlanname = obj.re_search_children(r"name")
        vlanname = vlanname[0].text.split()[-1] if vlanname else "No Name!"
        if not usedby and 'y' in input(
                "VLAN " + vlan + " (" + vlanname +
                ") is not used by any interface. Remove?[y|N]: ").lower():
            continue
        else:
            vlans.append(vlan)
        if genconfig:
            newconfig.prepend_line("!")
            for child in obj.children:
                newconfig.prepend_line(child.text)
            newconfig.prepend_line(obj.text)
            # newconfig.append_line("!")
            # newconfig.append_line(obj.text)
            # for child in obj.children:
            #     newconfig.append_line(child.text)
    if genconfig:
        newconfig.commit()
    return natsorted(vlans, key=lambda x: x.lower())


def add_snooping(newconfig, vlans):
    """
    Add DHCP snooping commands to new configuration file

    :param newconfig: CiscoConfParse object, representing the "new" config file
    :param vlans: List of VLANs to add
    """
    newconfig.append_line("!")
    newconfig.append_line("!")
    newconfig.append_line("ip dhcp snooping vlan " + ",".join(vlans))
    newconfig.append_line("no ip dhcp snooping information option")
    newconfig.append_line("ip dhcp snooping")
    newconfig.append_line("vlan internal allocation policy ascending")
    newconfig.commit()


def migrate_ports(oldconfig, newconfig, hostname, switch_type):
    """
    Map and transfer configuration settings of old ports

    Searches for Excel workbooks in the `./cutsheer_dir/` directory.
    The "cutsheet" files are generated by TurboClerk which are pulled from
    Netdoc.

    As of May 2016, Carlos Bassett has set a standard for worksheet layouts and
    file names, as follows:

    * The spreadsheets must have each tab nammed from the source switch and the
      first column *MUST* be the port name.
    * Any jack associated to this port must be in the row, otherwise
      configuration will not be transferred.
    * The file with the current port-jack mappings must have the building code
      in the name, along with "as is".
    * The file with the future port-jack mappings must have the building code
      in the name, along with "to be".
    * The file with the future port-jack mappings *MUST* begin listing ports in
      the third row.

    Note that this function has the potential to break if corresponding jacks
    are found from a different existing switch.
    A future workaround of loading the switch it is found from has been added
    as a TODO

    .. todo:: Remove `print` statements when `interfaces_for_review` is
              completed

    All returned variables are intended for printing out
    warnings/notices/debugging

    :param oldconfig: CiscoConfParse object of existing configuration file
    :param newconfig: CiscoConfParse object, representing the "new" config file
    :param hostname: Name of the new switch
    :param switch_type: the index of switch_models that represents the "new"
                        switch model

    :returns: _blades_ -- Detected blade numbers in the stack
              _nojacks_ -- Port names on the new switch that do not have a jack
              associated with them
              _newjacks_ -- Port names on the new switch that have jacks
              associated to them, however they do not exist in any As-Is
              spreadsheets
    :rtype: (Set, List, List)
    """
    # TODO: Detect jacks originating from a different source switch and
    # temporarily load their configuration file
    hostnotfound = True
    warnings.simplefilter("ignore")
    # Supervisor blade is not added to spreadsheets
    blades = set()
    nojacks = []
    newjacks = []
    oldhost = oldconfig.find_lines(r"^hostname")[0].split()[1].lower()
    allconfigfiles = [
        x for x in os.listdir(config_dir) if (
            "confg" in x.lower())]
    otherconfigs = {}
    # Filter by building
    filtered_files = []
    while True:
        filtered_files = [x for x in os.listdir(cutsheet_dir) if
                          (x.split()[0].split("-")[0].lower() in
                           hostname.lower()) and
                          ("as-is" in x.lower() or "as is" in x.lower())]
        if not filtered_files:
            no_files_found(cutsheet_dir)
        else:
            break
    jacks = {}
    for file in filtered_files:
        wb = openpyxl.load_workbook(cutsheet_dir + file)
        for ws in wb:
            try:
                sheet = ws.title.lower()
                alt = [m.span() for m in re.finditer(r"\d+$", sheet)]
                match = [x for x in allconfigfiles if sheet in x.lower()]
                if not match:
                    match = [
                        x for x in allconfigfiles if alt and sheet[
                            0:alt[0][0]] in x.lower()]
                if match and sheet not in otherconfigs.keys():
                    otherconfigs[sheet] = CiscoConfParse(
                        config_dir + match[0], factory=True)
            except IOError as e:
                print(e)
            for row in ws.rows:
                for cell in row:
                    # CiscoConfParse cannot accept leading 0"s in port names,
                    # ex. Fa1/0/01. While this is a localized issue of using
                    # Excel"s autofill, we"ll add the courtesy of dropping it
                    # from the port (not module) number ourselves
                    try:
                        portname = row[0].value
                        if '/' not in portname:
                            break  # Ensure the row is for a port
                        indices = [m.span()
                                   for m in re.finditer(r"\d+", portname)]
                        portname = (portname[0: indices[0][0]] +
                                    "/".join([str(int(portname[num[0]:num[1]]))
                                              for num in indices]))
                        if int(cell.value) in jacks.keys():
                            print(
                                "WARNING! DUPLICATE JACK FOUND IN AS-IS:",
                                cell.value)
                            print("SWITCH:", ws.title.lower())
                            print("PORT:", portname)
                        jacks[int(cell.value)] = {
                            "old port": portname,
                            "old switch": ws.title.lower()
                        }
                    # except Exception as e:
                        # print(e)
                    except (TypeError, ValueError):
                        continue
    while True:
        filtered_files = [x for x in os.listdir(cutsheet_dir) if
                          (x.split()[0].split("-")[0].lower() in
                           hostname.lower()) and
                          ("to-be" in x.lower() or "to be" in x.lower())]
        if not filtered_files:
            no_files_found(cutsheet_dir)
        else:
            break
    while True:
        for file in filtered_files:
            wb = openpyxl.load_workbook(cutsheet_dir + file)
            for ws in wb:
                if not hostname.lower() in ws["A1"].value.lower():
                    continue  # This isn"t a worksheet for the switch
                hostnotfound = False
                for row in ws.rows:
                    try:
                        portname = row[0].value
                        if '/' not in portname:
                            continue  # Ensure the row is for a port
                        indices = [m.span()
                                   for m in re.finditer(r"\d+", portname)]
                        portname = (portname[0:indices[0][0]] +
                                    "/".join([str(int(portname[num[0]:num[1]]))
                                              for num in indices]))
                        # This grabs the blade number. re.finditer will look
                        # for all digits and will be stored in 'm'. m.span will
                        # give the index numbers for where the match starts
                        # (inclusive) and where it ends (exclusive), and
                        # returns them as tuples. Since the blade is the first
                        # number, we want the first match. We then get the
                        # first number from that tuple since it's the index
                        # number of the string. You may be asking, "wait,
                        # aren't the cutsheets configured to only use the first
                        # letter of the port type, e.g. 'F|G|T'?" Yes, that's
                        # how it is currently, but since the cutsheet generator
                        # was designed by a student and edited by students, I
                        # have this unwarranted fear of someone changing 'T' to
                        # 'Te', etc. after I am long gone and this is still in
                        # use
                        blades.add(
                            int(portname[
                                [m.span()
                                 for m in re.finditer(r"\d", portname)][0][0]
                            ])
                        )
                        transferred = False
                        for cell in row:
                            try:
                                # Does the host match the one from our config?
                                if oldhost in jacks[int(cell.value)][
                                        "old switch"]:
                                    port = oldconfig.find_interface_objects(
                                        jacks[int(cell.value)]["old port"])
                                    newconfig.append_line("!")
                                    newconfig.append_line(
                                        "interface " + portname)
                                    for child in port[0].children:
                                        newconfig.append_line(child.text)
                                    transferred = True
                                    break  # Got the jack, move to next row
                                # If not, try loading it
                                else:
                                    host = jacks[int(cell.value)]["old switch"]
                                    cfg = (otherconfigs[host] if host in
                                           otherconfigs.keys() else None)
                                    if cfg:
                                        port = cfg.find_interface_objects(
                                            jacks[int(cell.value)]["old port"])
                                        newconfig.append_line("!")
                                        newconfig.append_line(
                                            "interface " + portname)
                                        for child in port[0].children:
                                            newconfig.append_line(child.text)
                                        transferred = True
                                        del cfg
                                        break  # Got the jack, move to next row
                                    else:
                                        raise IOError(
                                            'Could not find config for', host)
                            except IOError as e:
                                print(e)
                            except IndexError:
                                print("Could not load configuration for",
                                      jacks[int(cell.value)]["old port"],
                                      "from",
                                      jacks[int(cell.value)]["old switch"])
                            # Failures for int() cast
                            except (ValueError, TypeError):
                                continue  # Try next cell
                            except KeyError:
                                # We have a valid jack number but it's new
                                newjacks.append(portname)
                                newconfig.append_line("!")
                                newconfig.append_line("interface " + portname)
                                newconfig.append_line(" %%NEW CONNECTION%%")
                                transferred = True
                                break
                        if not transferred:
                            nojacks.append(portname)
                            newconfig.append_line("!")
                            newconfig.append_line("interface " + portname)
                            newconfig.append_line(" shutdown")
                    except TypeError:
                        continue
            del wb
        if hostnotfound:
            print(
                "A To-be cutsheet for the switch",
                hostname,
                "was not found!")
            print(
                ("If you need to modify the workbooks, make the changes "
                    "before responding."))
            if "n" in force_user_input(
                    "Would you like to retry? ",
                    r"[Yy][EeSs]*|[Nn][Oo]?").lower():
                sys.exit("Aborting...")
        else:
            break
    newconfig.commit()
    del otherconfigs
    if switch_models[switch_type] == "4506":
        blades = set(xrange(0, 7))  # Blades CAN be left empty, so there's that
    return blades, nojacks, newjacks


def interfaces_for_review(newconfig, nojacks, newjacks):
    """
    Searches for interfaces on the "new" device with non-standard configs to
    be reviewed manually.

    Searches for statically set PoE, duplex, operating speed, no defined
    switchport mode

    :param newconfig: CiscoConfParse object representing the "new"
                      configuration file
    """
    power = newconfig.find_objects_w_child(r"^interf", r"^ power")
    # print power
    if power:
        print(
            "\nManually review the following Interfaces for Power settings:")
        power_ports = []
        for obj in power:
            # print(" ", obj.text)
            power_ports.append(obj.text.split()[-1])
        print(condensify_ports(power_ports))

    duplex = newconfig.find_objects_w_child(r"^interf", r"^ duplex")
    if duplex:
        print(
            "\nManually review the following Interfaces for Duplex settings:")
        duplex_ports = []
        for obj in duplex:
            # print(" ", obj.text)
            duplex_ports.append(obj.text.split()[-1])
        print(condensify_ports(duplex_ports))

    speed = newconfig.find_objects_w_child(r"^interf", r"^ speed")
    if speed:
        print("\nManually review the following Interfaces for Speed settings:")
        speed_ports = []
        for obj in speed:
            # print(" ", obj.text)
            speed_ports.append(obj.text.split()[-1])
        print(condensify_ports(speed_ports))

    access = newconfig.find_objects_wo_child(
        r"^interf", r"^ switchport mode")
    access = [obj for obj in access if not obj.re_search_children(
        r"^[\s]?shut(down)?$")]
    if access:
        access_ports = []
        for obj in access:
            # print(" ", obj.text)
            access_ports.append(obj.text.split()[-1])
        print(
            ("\nThese interfaces did not specify Access or Trunk mode and "
                "were set to Access by default.\nManually review the "
                "following:"))
        print(condensify_ports(access_ports))
        for obj in access:
            obj.append_to_family(" switchport mode access")

    if nojacks:
        print(
            ("\nThe following interfaces did not have any jacks specified and"
             " will be shut down:"))
        print(condensify_ports(nojacks))

    if newjacks:
        print(
            ("\nThe following interfaces appear to have new jacks associated. "
             "Please manually configure them:"))
        print(condensify_ports(newjacks))
    print("")
    newconfig.commit()


def add_voice_vlan(voicevlan, newconfig):
    """
    Add voice VLAN to access ports that do not have it

    .. todo:: Remove `print` statements when `interfaces_for_review` is

    :param voicevlan: a VLAN represented as a string or int
    :param newconfig: CiscoConfParse object representing the "new"
                      configuration file
    """
    modified = []
    Voice = " switchport voice vlan " + str(voicevlan)  # Cast, just in case
    NewIntfs = newconfig.find_objects_wo_child(
        r"^interf", r"^ switchport mode trunk")
    for obj in NewIntfs:
        # testing to see if access port, should probably adjust to search for
        # switchport mode access, but some ports have this explicit and some
        # don't
        if obj.access_vlan != 0:
            hasVoice = False
            for child in obj.children:
                if "voice" in child.text:
                    hasVoice = True
            if not hasVoice:
                obj.append_to_family(Voice)
                modified.append(obj.text)
    # modified = []
    if len(modified) > 0:
        print(
            ("The following ports had a Voice VLAN added. Please manually "
                " check that this is appropriate:"))
        for each in modified:
            print(" ", each)
    newconfig.commit()
    return modified


def trunk_cleanup(newconfig):
    """
    Remove access mode configuration on trunk ports

    Removes `access/voice vlan` configs, `spanning-tree portfast`, and
    `no snmp trap link-status`

    .. todo:: Detect and remove VLANs from VLAN ranges

    :param newconfig: CiscoConfParse object representing the "new"
                      configuration file
    """
    #
    TrunkInts = newconfig.find_objects_w_child(
        r"^interf", r"^ switchport mode trunk|shutdown")
    for obj in TrunkInts:
        for child in obj.children:
            if r"switchport access" in child.text:
                child.delete()
            elif r"switchport voice" in child.text:
                child.delete()
            elif r"spanning-tree portfast" in child.text:
                child.delete()
            elif r"no snmp" in child.text:
                child.delete()
    newconfig.commit()
    # Now we want to remove from trunk ports VLANs 1,1002,1003,1004,1005
    TrunkInts = newconfig.find_objects_w_child(
        r"^interf", r"^ switchport mode trunk")
    for obj in TrunkInts:
        TrunkLines = obj.re_search_children(r"switchport trunk allowed vlan")
        # obj.delete_children_matching(r"switchport trunk allowed vlan")
        for child in TrunkLines:
            # TrunkVlans = []
            Add = False
            if "add" in child.text.lower():
                Add = True
            TrunkNumbers = set(child.text.split()[-1].split(","))
            NumsToRemove = set(
                ["1", "1002-1005", "1002", "1003", "1004", "1005"])
            TrunkNumbers = natsorted(
                list(
                    TrunkNumbers -
                    NumsToRemove),
                lambda x: x.lower())
            if len(TrunkNumbers) > 0:
                if Add:
                    TrunkConfigLine = " switchport trunk allowed vlan add " + \
                        ",".join(TrunkNumbers)
                    child.replace(
                        r"^ switchport trunk allowed vlan add .*",
                        TrunkConfigLine)
                else:
                    TrunkConfigLine = " switchport trunk allowed vlan " + \
                        ",".join(TrunkNumbers)
                    child.replace(
                        r"^ (?!add)switchport trunk allowed vlan .*",
                        TrunkConfigLine)
                # obj.append_to_family(TrunkConfigLine)
    newconfig.commit()


def remove_mdix_and_dot1q(newconfig):
    """
    Remove MDIX and dot1q from all interfaces

    .. note:: Should be run after `trunk_cleanup()`

    Keyword arguments:
    newconfig -- CiscoConfParse object, representing the "new" config file
    """
    MdixInts = newconfig.find_objects("no\smdix\sauto")
    for obj in MdixInts:
        if obj != []:
            obj.delete()
    newconfig.commit()
    Dot1qInts = newconfig.find_objects(
        "switchport\strunk\sencapsulation\sdot1q")
    for obj in Dot1qInts:
        if obj != []:
            obj.delete()
    newconfig.commit()


def access_cleanup(newconfig):
    """
    Remove trunk configuration for all ports set to access mode

    :param oldconfig: CiscoConfParse object of existing configuration file
    :param newconfig: CiscoConfParse object, representing the "new" config file
    """
    AccessInts = newconfig.find_objects_w_child(
        r"^interf", r"^ switchport mode access")
    for obj in AccessInts:
        for child in obj.children:
            if r"switchport trunk" in child.text:
                child.delete()
            child.replace(
                r"spanning\-tree\sportfast\sedge.*",
                "spanning-tree portfast")
    newconfig.commit()
    # I found that some configs are missing the "mode access", so I am also
    # applying to any that are missing "switchport mode"
    # AccessInts = newconfig.find_objects_wo_child(
    #     r"^interf", r"^ switchport mode")
    # for obj in AccessInts:
    #     for child in obj.children:
    #         if r"switchport trunk" in child.text:
    #             child.delete()
    #         child.replace_children(
    #                                r"spanning\-tree\sportfast\sedge.*",
    #                                "spanning-tree portfast")
    # newconfig.commit()


def extract_management(oldconfig, newconfig):
    """
    Extract the management VLAN and add it to new config file

    Assuming the target equipment is a layer 2 switch with only one management
    VLAN, the VLAN config is extracted and the option to retain IP information
    is provided.
    `ip tacacs source-interface <VLAN>` is added but only necessary for a 4506;
    this command will be ignored on all other models.

    :param oldconfig: CiscoConfParse object of existing configuration file
    :param newconfig: CiscoConfParse object, representing the "new" config file
    """
    Vlan1 = oldconfig.find_objects(r"^interface Vlan1$")
    for vlan in Vlan1:
        vlan.delete()
    oldconfig.commit()
    VlanInts = oldconfig.find_objects_wo_child(
        r"^interface Vlan", r"^ ip address")
    for vlan in VlanInts:
        newconfig.append_line(vlan.text)
        for child in vlan.children:
            newconfig.append_line(child.text)
        newconfig.append_line("!")
    newconfig.commit()
    # TODO: Prompt user for any additional configuration to include
    ManagementVlan = oldconfig.find_objects_w_child(
        r"^interface Vlan3\d{2}", r"^ ip address")
    # print(ManagementVlan)
    if not ManagementVlan:
        print("No management VLAN defined!")
        print("Please verify and manually fix later.")
        # if not 'n' in input("Would you like to create one?[Y|n]: ").lower():
        #    vnum = force_user_input("VLAN number: ",r"3\d{2}")
    else:
        if len(ManagementVlan) > 1:
            print(
                ("Several VLANs were found with IP configuration. (Is this "
                    "from a router?)"))
            print("The following were found:")
            ok = False
            while not ok:
                for option, vlan in zip(
                        xrange(1, len(ManagementVlan) + 1), ManagementVlan):
                    print(' [' + str(option) + ']', vlan.text)
                choice = None
                while choice not in xrange(1, len(ManagementVlan) + 1):
                    try:
                        choice = int(
                            input(
                                ("Please select a source VLAN for "
                                    "management: ")))
                    except:
                        pass
                # Remember: We're enumerating from 1, so cut the index back
                choice -= 1
                print(ManagementVlan[choice].text)
                for child in ManagementVlan[choice].children:
                    print(child.text)
                ok = True if 'y' in input(
                    "Is this correct?[y|N]: ").lower() else False
        else:
            choice = 0
        mgmt = ManagementVlan[choice]
        newconfig.append_line("!")
        newconfig.append_line(mgmt.text)
        if "n" in input(
                "Will this equipment use the same IP address?[Y|n]: ").lower():
            while True:
                newip = force_user_input(
                    "Enter new IP address: ",
                    r'(\d{1,3}\.){3}\d{1,3}')
                newgate = force_user_input(
                    "Enter new Gateway: ",
                    r'(\d{1,3}\.){3}\d{1,3}')
                newmask = force_user_input(
                    "Enter new subnet Mask: ",
                    r'(\d{1,3}\.){3}\d{1,3}')
                print(" IP     :", newip)
                print(" Gateway:", newgate)
                print(" Mask   :", newmask)
                if is_ip(newip) and is_ip(newgate) and is_ip(newmask):
                    if 'y' in force_user_input(
                            "Is this correct? ",
                            r"[Yy][EeSs]*|[Nn][Oo]?").lower():
                        break
                else:
                    print("One of the addresses is not a valid IP format!")
            newconfig.append_line(" ip address " + newip + " " + newmask)
            newconfig.append_line(" no ip route-cache")
            newconfig.append_line(" no ip mroute-cache")
            newconfig.append_line(" no shutdown")
            newconfig.append_line("!")
            newconfig.append_line("ip default-gateway " + newgate)
            newconfig.append_line("!")
        else:
            for child in mgmt.children:
                if 'ip address' in child.text:
                    newconfig.append_line(child.text)
            newconfig.append_line(" no ip route-cache")
            newconfig.append_line(" no ip mroute-cache")
            newconfig.append_line(" no shutdown")
            newconfig.append_line("!")
            newconfig.append_line("!")
            DefaultGateway = oldconfig.find_objects(r"^ip default-gateway")
            if len(DefaultGateway) > 1:
                print("Several default gateway commands discovered (WHAT?!)")
                print("The following were found:")
                while True:
                    for option, gate in zip(
                        xrange(1, len(DefaultGateway) + 1),
                            DefaultGateway):
                        print(' [' + str(option) + ']', gate.text)
                    choice = None
                    while choice not in xrange(1, len(DefaultGateway) + 1):
                        try:
                            choice = int(
                                force_user_input(
                                    "Please select the default gateway: "))
                        except:
                            pass
                    # Remember: We're enumerating from 1, so cut the index back
                    choice -= 1
                    print(DefaultGateway[choice].text)
                    if 'y' in input("Is this correct?[y|N]: ").lower():
                        break
                newconfig.append_line(DefaultGateway[choice].text)
            elif len(DefaultGateway) == 0:
                while True:
                    newgate = force_user_input(
                        "No default gateway found. Please type it here: ")
                    if is_ip(newgate):
                        newconfig.append_line("ip default-gateway " + newgate)
                        break
            else:
                newconfig.append_line(DefaultGateway[0].text)
            newconfig.append_line("!")
        # Only the 4506 uses the following line, but it will be ignored on
        #  other devices anyways, so we"ll just leave it here
        newconfig.append_line("ip tacacs source-interface " +
                              mgmt.text.split()[-1])
        newconfig.commit()


def file_export(outputfile, newconfig):
    """
    Save current configuration to a file

    Exports to the directory defined by internal/global var 'output_dir'

    :param outputfile: Desired file name
    :param newconfig: CiscoConfParse object, representing the "new" config file
    """
    try:
        newconfig.save_as(output_dir + outputfile)
        print("\nFile saved as", outputfile)
    except IOError as e:
        print("Unable to save file:")
        print(e)


def setup_feeds(newconfig, switch_type, blades, vlans):
    """
    Configure feedports

    Allows the user to define as many feedports as desired.
    Checks the validity of the port name as defined by a regex string.

    :param newconfig: CiscoConfParse object, representing the "new" config file
    :param switch_type: The representation of the switch model in the form of
                        the switch_models index
    :param blades: A Set of all blade numbers in the stack
    :param vlans: -- A List of all VLANs transferred to the new configuration
                     file
    """
    if "y" in input(
            "\nWould you like to set up feedports?[y|N]: ").lower():
        try:
            print("When you are finished, type 'no' for the feedport name")
            print("(Use 'T' or 'G' instead of 'Ten' or 'Gig'!)")
            while True:
                feedport = input("Feedport name: ").upper()
                if feedport and not feedport.lower() == "no":
                    blade = int(feedport[
                        [m.span() for m in re.finditer(r"\d", feedport)][0][0]
                    ])
                    if re.search(feed_ports_regex[switch_models[switch_type]],
                                 feedport) and (blade in blades):
                        exists = newconfig.find_objects(
                            r"^interface " + feedport.strip())
                        if len(exists) > 1:
                            print("Uh, your interface matches several others:")
                            for each in exists:
                                print(" ", each.text)
                            print("Abandoning changes...")
                        elif len(exists) == 1:
                            existing = exists[0]
                            print("Interface already exists!")
                            print(existing.text)
                            for line in existing.children:
                                print(line.text)
                            if "y" in input("Overwrite?[y|N]: ").lower():
                                for child in existing.children:
                                    child.delete()
                                newconfig.commit()
                                _setup_feed(newconfig, feedport, switch_type)
                        else:
                            _setup_feed(newconfig, feedport, switch_type)
                    else:
                        print(
                            ("That is an invalid port. Either the spelling is "
                                "wrong or the numbering is out of range!"))
                        if 'y' in input(
                                ("Would you like to use it "
                                    "anyways?[y|N]: ")).lower():
                            exists = newconfig.find_objects(
                                r"^interface " + feedport.strip())
                            if len(exists) > 1:
                                print(
                                    ("Uh, your interface matches several "
                                     "others:"))
                                for each in exists:
                                    print(" ", each.text)
                                print("Abandoning changes...")
                            elif len(exists) == 1:
                                existing = exists[0]
                                print("Interface already exists!")
                                print(existing.text)
                                for line in existing.children:
                                    print(line.text)
                                if "y" in input("Overwrite?[y|N]: ").lower():
                                    for child in existing.children:
                                        child.delete()
                                    newconfig.commit()
                                    _setup_feed(
                                        newconfig, feedport, switch_type)
                            else:
                                _setup_feed(newconfig, feedport, switch_type)
                else:
                    break
        except Exception as e:
            print("Exception:", e)


def _setup_feed(newconfig, feedport, switch_type):
    """
    Create or modify the desired feed port

    Only to be called by the setup_feeds function.
    Prompts for description, will add 'dot1q' for 3560 models
    Puts port into trunk mode, adds all VLANs, adds DHCP snooping trust

    :param newconfig: CiscoConfParse object, representing the "new" config file
    :param feedport: String representing the desired feed name
    :param switch_type: Representation of the switch model in the form of the
                        switch_models index
    """
    existing = newconfig.find_objects(r"^interface " + feedport.strip())
    if existing:
        exists = existing[0]
        if "y" in input("Would you like to add a description?[y|N]: ").lower():
            exists.append_to_family(" description " + input("Description: "))
        if switch_models[switch_type] == "3560":
            exists.append_to_family(" switchport trunk encapsultion dot1q")
        exists.append_to_family(" switchport mode trunk")
        exists.append_to_family(
            " switchport trunk allowed vlan " +
            ",".join(vlans))
        exists.append_to_family(" ip dhcp snooping trust")
    # elif newconfig.has_line_with(r"^hostname"):
    #     newconfig.insert_after(
    #         r"^hostname",
    #         " ip dhcp snooping trust")
    #     newconfig.insert_after(
    #         r"^hostname",
    #         " switchport trunk allowed vlan " +
    #         ",".join(vlans))
    #     newconfig.insert_after(
    #         r"^hostname",
    #         " switchport mode trunk")
    #     if switch_models[switch_type] == "3560":
    #         newconfig.insert_after(
    #             r"^hostname",
    #             " switchport trunk encapsultion dot1q")
    #     if "y" in input(
    #                     ("Would you like to add a description?[y|N]: ")
    #                    ).lower():
    #         newconfig.insert_after(
    #             r"^hostname",
    #             " description " +
    #             input("Description: "))
    #     newconfig.insert_after(
    #         r"^hostname",
    #         "interface " + feedport)
    #     newconfig.insert_after(r"^hostname", "!")
    else:
        newconfig.append_line("!")
        newconfig.append_line("interface " + feedport)
        if "y" in input("Would you like to add a description?[y|N]: ").lower():
            newconfig.append_line(" description " + input("Description: "))
        if switch_models[switch_type] == "3560":
            newconfig.append_line(" switchport trunk encapsultion dot1q")
        newconfig.append_line(" switchport mode trunk")
        newconfig.append_line(
            " switchport trunk allowed vlan " +
            ",".join(vlans))
        newconfig.append_line(" ip dhcp snooping trust")
        newconfig.append_line("!")
    newconfig.commit()


def get_switch_model():
    """Prompt user to select model from compatible list

    :returns: The user's input as the internal `switch_models` index
    :rtype: Int
    """
    print("Compatible switch models:")
    for switch in enumerate(switch_models):
        print(" [" + str(switch[0] + 1) + "]", switch[1])
    switch_type = None
    while switch_type not in xrange(0, len(switch_models)):
        switch_type = int(
            input("What switch model are you programming? ")) - 1
    return switch_type


def get_configs():
    """
    Set the configuration file to pull data from
    Prompts user for file name

    :returns: oldconfig -- Existing/source configuration file
              newconfig -- Container for the new device's configuration
    :rtype: CiscoConfParse, CiscoConfParse
    """
    directory = sorted([x for x in os.listdir(config_dir) if "confg" in x])
    if not directory:
        no_files_found(config_dir)
    for item in enumerate(directory):
        print(" [" + str(item[0] + 1) + "]", item[1])
    name = None
    while name not in xrange(1, len(directory) + 1):
        name = int(
            input("Please select the file number of the old config: "))
    filename = directory[name - 1]
    oldconfig = CiscoConfParse(config_dir + filename, factory=True)
    # # The new parser needs a file associated with it, so create a throwaway.
    # # Trying to give it a legit file strangely invokes an error later on...
    if sys.platform.startswith('win'):
        newconfig = CiscoConfParse(
            tempfile.NamedTemporaryFile(
                dir=output_dir).file, factory=True)
    else:
        newconfig = CiscoConfParse(os.tmpfile(), factory=True)
    return oldconfig, newconfig


def force_user_input(display, expect=r""):
    """
    Enforce that the user input at least one character

    :param display: String to display as the input prompt
    :param expect: Regex string representing the required format of the
                   response before returning to caller. Defaults to an empty
                   string (match any)

    :returns: User's input
    :rtype: String
    """
    name = ""
    ok = False
    while len(name) == 0 and not ok:
        name = input(display)
        if re.match(expect, name):
            ok = True
        else:
            name = ""
    return name


def set_voice_vlan(oldconfig):
    """
    Select and add a voice VLAN to add to access ports

    :param oldconfig: CiscoConfParse object of existing configuration file
    """
    print("Available voice VLANs:")
    vlans = oldconfig.find_objects(r"^vlan 1\d\d$")
    skipvoice = False
    if len(vlans) > 1:
        voicevlan = ""
        while len(voicevlan) == 0:
            for v, num in zip(vlans, xrange(1, len(vlans) + 1)):
                print(" [" + str(num) + "]", v.text.split()
                      [-1], "-", v.children[0].text.split()[1:])
            print(" [0] Skip")
            voicevlan = input("    Enter voice VLAN: ")
            try:
                if int(voicevlan) == 0:
                    skipvoice = True
                    break
                else:
                    voicevlan = vlans[int(voicevlan) - 1].text.split()[-1]
            except:
                voicevlan = ""  # Invalid input
    elif len(vlans) == 1:
        if "n" in input("Only one voice VLAN found: " +
                        vlans[0].text.split()[-1] + " - " +
                        str(vlans[0].children[0].text.split()[1:]) +
                        ". Use this?[Y|n]: ").lower():
            skipvoice = True
        else:
            voicevlan = vlans[0].text.split()[-1]
    else:
        print("No standard voice VLAN detected.")
        skipvoice = True
    if not skipvoice:
        add_voice_vlan(voicevlan, newconfig)


def no_files_found(directory):
    """
    Allows the user to move files to correct directory or exit early

    :param directory: The folder in which the files should be located
    """
    print("Files not found in directory:", os.path.abspath(directory))
    print("If you would like to add/move them, please do so now!")
    if 'n' not in input("Would you like to retry?[Y|n] ").lower():
        return
    else:
        sys.exit("Exiting configuration generator")


def is_ip(addr):
    return True if re.match(r'(\d{1,3}\.){3}\d{1,3}', addr.strip()) else False

    # This module is designed to allow running it from a CLI or importing it
    # and calling its functions, however it revolves mostly around Layer 2
    # device configuration
if __name__ == "__main__":
    oldconfig, newconfig = get_configs()
    switch_type = get_switch_model()
    hostname = force_user_input("Enter hostname of new switch: ").upper()
    outputfile = input(
        "Enter output file name (default - " + hostname + ".txt): ")
    outputfile = outputfile if outputfile else hostname + ".txt"
    createconfig = ("n" not in input(
        "Generate full config file?[Y|n]: ").lower())

    blades, nojacks, newjacks = migrate_ports(
        oldconfig, newconfig, hostname, switch_type)
    # Add ip dhcp snooping later: adding it immediately after interfaces
    # causes a bug if trying to use file as startup-config
    vlans = vlan_extract(oldconfig, newconfig, feed_ports_regex[
        switch_models[switch_type]], createconfig)
    setup_feeds(newconfig, switch_type, blades, vlans)
    interfaces_for_review(newconfig, nojacks, newjacks)
    set_voice_vlan(oldconfig)
    access_cleanup(newconfig)
    trunk_cleanup(newconfig)

    # This must be run AFTER trunk_cleanup()
    if not switch_models[switch_type] == "3560":
        remove_mdix_and_dot1q(newconfig)
    if createconfig:
        baseconfig = ".txt"
        if (switch_type == len(switch_models) - 1):
            baseconfig = "baseconfig.txt"
        else:
            baseconfig = switch_models[switch_type] + "base.txt"
        newconfig.prepend_line("!")
        newconfig.prepend_line("hostname " + hostname)
        newconfig.prepend_line("!")
        newconfig.commit()
        # vlans = prune_unused_vlans(
        #     newconfig, vlans, feed_ports_regex[
        #         switch_models[switch_type]])
        extract_management(oldconfig, newconfig)
        add_snooping(newconfig, vlans)
        newconfig.append_line("!")
        with open(template_dir + baseconfig, "r") as b:
            for line in b:
                newconfig.append_line(line.rstrip())
        newconfig.commit()

    file_export(outputfile, newconfig)
